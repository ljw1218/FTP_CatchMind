import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QLabel, QLineEdit, QTextEdit, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox
from PyQt5.QtCore import Qt,QPoint
import pymysql,socket
from PyQt5.QtWidgets import QWidget,QTableWidgetItem,QTableWidget,QColorDialog
from PyQt5.QtGui import QPainter, QPen, QColor,QFont, QPixmap
import ftplib,os,shutil,random
from openpyxl import load_workbook

class MainUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Main')
        self.setGeometry(100, 100, 600, 400)
        self.conn = pymysql.connect(host='host',port=3306,user='root', passwd='1234',db='test',charset='utf8',autocommit=True)
        self.MariaDB = self.conn.cursor()
        self.selected_index = 1
        self.register_ip()
        self.Init()

        # 왼쪽 리스트 뷰
        self.list_view = QTableWidget()
        self.list_view.setFixedWidth(381)
        self.list_view.setColumnCount(3)
        self.list_view.setHorizontalHeaderLabels(['Index', '등록자', '정답자 수'])
        self.load_data_to_list_view()
        self.list_view.cellClicked.connect(self.on_list_item_clicked)

        # 새로고침 버튼
        refresh_button = QPushButton('새로고침')
        refresh_button.clicked.connect(self.load_data_to_list_view)

        # 사용자 닉네임 라벨
        query = 'SELECT Nickname FROM userlist where IP = "%s"'%self.ip
        self.MariaDB.execute(query)
        rst = self.MariaDB.fetchone()
        self.nickname_label = QLabel('닉네임 : %s'%rst[0])

        # 사용자 닉네임 변경 버튼
        nickname_change_button = QPushButton('닉네임 변경')
        nickname_change_button.clicked.connect(self.open_nickname_dialog)

        # 정답 갯수 라벨
        self.answer_count_label = QLabel('정답 수')

        # 문제 등록 버튼
        question_register_button = QPushButton('문제 등록')
        question_register_button.clicked.connect(self.open_question_dialog)

        question_solve_button = QPushButton('문제 풀기')
        question_solve_button.clicked.connect(self.open_solve_window)

        # 메인
        layout = QHBoxLayout()
        layout.addWidget(self.list_view)
        layout_widget = QWidget()
        layout_widget.setLayout(layout)

        # 우측
        right_layout = QVBoxLayout()
        right_layout.addWidget(refresh_button)
        right_layout.addWidget(self.nickname_label)
        right_layout.addWidget(nickname_change_button)
        right_layout.addWidget(self.answer_count_label)
        right_layout.addWidget(question_register_button)
        right_layout.addWidget(question_solve_button)
        right_layout.addStretch()

        # 전체
        main_layout = QHBoxLayout()
        main_layout.addWidget(layout_widget)
        main_layout.addLayout(right_layout)

        central_widget = QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def Init(self):
        if not os.path.exists('./img'):
            os.makedirs('./img')
        if not os.path.exists('./img/word.xlsx'):
            with ftplib.FTP() as ftp:
                # ftp = FTP('ftp://59.19.145.3:25904/Test/img')
                ftp.connect(host='host', port=25904)
                ftp.encoding = 'utf-8'
                username = 'skywave'
                password = '1234'
                remote_path = './img'
                filename = 'word.xlsx'
                ftp.login(user=username, passwd=password)
                ftp.cwd('/Test/img')
                file_path = remote_path + '/' + filename

                with open(filename, 'wb') as file:
                    ftp.retrbinary(f"RETR %s" % filename, file.write)
                    file.close()
                ftp.quit()
                shutil.move("%s" % filename, "%s" % file_path)

    def register_ip(self):
        ip = self.get_local_ip()
        self.ip = ip

        query = "SELECT * FROM userlist WHERE IP = '%s'"%ip
        self.MariaDB.execute(query)
        result = self.MariaDB.fetchone()

        if result is None:
            nickname = ip.split('.')[-1]  
            query = "INSERT INTO userlist (IP, Nickname, cnt, crr_list) VALUES ('%s', '%s', %d, '[]')"%(ip,nickname,0)
            self.MariaDB.execute(query)

    def load_data_to_list_view(self):
        self.list_view.clear()
        self.list_view.setHorizontalHeaderLabels(['Index', '등록자', '정답자 수'])

        query = "SELECT QIndex, username, ans_list FROM question"
        self.MariaDB.execute(query)
        results = self.MariaDB.fetchall()
        i = 0
        self.list_view.setRowCount(len(results))
        for row in results:
            index = row[0]
            registrant = row[1]
            namequery = 'SELECT NickName from userlist where IP = "%s"'%row[1]
            self.MariaDB.execute(namequery)
            rst = self.MariaDB.fetchone()
            reg = rst[0]
            answer_list = row[2].split(',')
            answer_count = len(answer_list) if answer_list[0] else 0

            '''item = QListWidgetItem()
            item.setTextAlignment(Qt.AlignCenter)
            item.setData(Qt.UserRole, index)  # 데이터에 Index 저장

            index_item = QTableWidgetItem(str(index))
            index_item.setTextAlignment(Qt.AlignCenter)

            registrant_item = QTableWidgetItem(registrant)
            registrant_item.setTextAlignment(Qt.AlignCenter)

            answer_count_item = QTableWidgetItem(str(answer_count))
            answer_count_item.setTextAlignment(Qt.AlignCenter)'''

            #self.list_view.addItem(item)
            self.list_view.setItem(i, 0, QTableWidgetItem((str)(index)))
            self.list_view.setItem(i, 1, QTableWidgetItem(reg))
            self.list_view.setItem(i, 2, QTableWidgetItem((str)(answer_count)))
            i = i + 1

    def on_list_item_clicked(self, row,column):
        index = self.list_view.item(row,0)
        self.selected_index = (int)(index.text())
        # TODO: B 테이블에서 해당 Index의 데이터 가져오기

    def open_solve_window(self):
        dialog = ProblemWindow(self.selected_index)
        dialog.exec_()

    def open_nickname_dialog(self):
        dialog = NicknameDialog(self)
        dialog.exec_()

    def open_question_dialog(self):
        dialog = QuestionDialog(self)
        dialog.exec_()

    def get_local_ip(self):
        host_name = socket.gethostname()
        ip_addr = socket.gethostbyname(host_name)

        return ip_addr


class NicknameDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('닉네임 변경')
        self.setFixedSize(320, 90)
        self.conn = pymysql.connect(host='host', port=3306, user='root', passwd='1234', db='test',
                                    charset='utf8', autocommit=True)
        self.MariaDB = self.conn.cursor()

        self.nickname_label = QLabel('닉네임')
        self.nickname_edit = QLineEdit()
        self.register_button = QPushButton('등록')
        self.cancel_button = QPushButton('취소')

        layout = QVBoxLayout()
        layout.addWidget(self.nickname_label)
        layout.addWidget(self.nickname_edit)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.register_button)
        button_layout.addWidget(self.cancel_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

        self.register_button.clicked.connect(self.register_nickname)
        self.cancel_button.clicked.connect(self.close)

    def register_nickname(self):
        ip = socket.gethostbyname(socket.gethostname())

        nickname = self.nickname_edit.text()
        if nickname:
            query = "UPDATE userlist SET Nickname = '%s' WHERE IP = '%s'"%(nickname,ip)
            self.MariaDB.execute(query)

            self.parent().nickname_label.setText(nickname)

        self.close()

class ProblemWindow(QDialog):
    def __init__(self, index,parent=None):
        super(ProblemWindow, self).__init__(parent)
        self.setWindowTitle("문제 풀기")
        self.setFixedSize(600, 460)
        self.conn = pymysql.connect(host='host', port=3306, user='root', passwd='1234', db='test',
                                    charset='utf8', autocommit=True)
        self.MariaDB = self.conn.cursor()

        self.database = None
        self.index = index
        self.answered = False
        self.ip = socket.gethostbyname(socket.gethostname())

        self.layout = QVBoxLayout()
        #self.layout.setContentsMargins(10, 10, 10, 10)
        self.central_widget = QWidget(self)
        self.central_widget.setLayout(self.layout)
        #self.setCentralWidget(self.central_widget)

        self.problem_label = QLabel()
        self.problem_label.setFont(QFont("Arial", 14))
        self.problem_label.setWordWrap(True)
        self.layout.addWidget(self.problem_label)

        self.image_label = QLabel()
        self.image_label.setFixedSize(570, 350)
        pixmap = self.load_ftp_img()
        self.image_label.setPixmap(pixmap)
        self.layout.addWidget(self.image_label)

        self.answer_edit = QLineEdit()
        self.layout.addWidget(self.answer_edit)

        button_layout = QHBoxLayout()
        submit_button = QPushButton("제출")
        submit_button.clicked.connect(self.submit_answer)
        button_layout.addWidget(submit_button)
        cancel_button = QPushButton("취소")
        cancel_button.clicked.connect(self.close)
        button_layout.addWidget(cancel_button)
        self.layout.addLayout(button_layout)

    def load_ftp_img(self):
        with ftplib.FTP() as ftp:
            #ftp = FTP('ftp://59.19.145.3:25904/Test/img')
            ftp.connect(host='59.19.145.3',port=25904)
            ftp.encoding = 'utf-8'
            username = 'skywave'
            password = '1234'
            remote_path = './img'
            filename = '%s.jpg'%self.index
            ftp.login(user=username,passwd=password)
            ftp.cwd('/Test/img')
            file_path = remote_path + '/' + filename

            with open(filename,'wb') as file:
                ftp.retrbinary(f"RETR %s"%filename,file.write)
                file.close()
            ftp.quit()
            shutil.move("%s"%filename,"%s"%file_path)
            pixmap = QPixmap('./img/%s'%filename)
            return pixmap

    def submit_answer(self):
        answer = self.answer_edit.text()
        if answer:
            if self.check_answer(self.index, answer):
                #self.add_answerer(self.index, self.ip)
                #if not self.answered:
                    #self.database.add_answer_index(self.ip, self.index)
                self.answered = True
                QMessageBox.information(self, "정답", "정답입니다!", QMessageBox.Ok)
            else:
                QMessageBox.warning(self, "오답", "오답입니다.", QMessageBox.Ok)
        else:
            QMessageBox.warning(self, "경고", "정답을 입력해주세요.", QMessageBox.Ok)

    def check_answer(self,index,answer):
        query = "SELECT ans FROM question WHERE QIndex = %d"%index
        self.MariaDB.execute(query)
        result = self.MariaDB.fetchone()
        if answer == result[0]:
            return True
        else:
            return False

class QuestionDialog(QDialog):
    class DrawingArea(QWidget):
        def __init__(self):
            super(QuestionDialog.DrawingArea, self).__init__()
            self.setFixedSize(570, 350)
            self.drawing = False
            self.last_point = QPoint()
            self.pixmap = QPixmap(self.size())
            self.pixmap.fill(Qt.white)
            self.color = QColor(0,0,0)

        def paintEvent(self, event):
            painter = QPainter(self)
            painter.drawPixmap(self.rect(), self.pixmap)

        def mousePressEvent(self, event):
            if event.button() == Qt.LeftButton:
                self.drawing = True
                self.erase = False
                self.last_point = event.pos()
            elif event.button() == Qt.RightButton:
                self.erase = True
                self.drawing = False
                self.last_point = event.pos()

        def mouseMoveEvent(self, event):
            if self.drawing:
                painter = QPainter(self.pixmap)
                painter.setPen(QPen(self.color, 2, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
                painter.drawLine(self.last_point, event.pos())
                painter.end()
                self.last_point = event.pos()
                self.update()
            elif self.erase:
                painter = QPainter(self.pixmap)
                painter.setPen(QPen(QColor(Qt.white),10, Qt.SolidLine, Qt.RoundCap, Qt.RoundJoin))
                painter.drawLine(self.last_point,event.pos())
                painter.end()
                self.last_point = event.pos()
                self.update()

        def mouseReleaseEvent(self, event):
            if event.button() == Qt.LeftButton and self.drawing:
                self.drawing = False
            elif event.button() == Qt.RightButton and self.erase:
                self.erase = False

        def change_color(self,color):
            self.color = color

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('문제 등록')
        self.setFixedSize(600, 430)
        self.conn = pymysql.connect(host='host', port=3306, user='root', passwd='1234', db='test',
                                    charset='utf8', autocommit=True)
        self.MariaDB = self.conn.cursor()

        self.drawing_area = self.DrawingArea()
        self.answer_label = QLabel('정답')
        self.answer_text_edit = QTextEdit()
        self.register_button = QPushButton('등록')
        self.cancel_button = QPushButton('취소')
        self.random_button = QPushButton('단어생성')
        self.color_button = QPushButton('색상변경')

        layout = QVBoxLayout()
        layout.addWidget(self.drawing_area)


        layout.addWidget(self.answer_label)
        layout.addWidget(self.answer_text_edit)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.register_button)
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.random_button)
        button_layout.addWidget(self.color_button)

        layout.addLayout(button_layout)
        self.setLayout(layout)

        self.register_button.clicked.connect(self.register_question)
        self.cancel_button.clicked.connect(self.close)
        self.random_button.clicked.connect(self.random_word)
        self.color_button.clicked.connect(self.color_change)

    def color_change(self):
        col = QColorDialog.getColor()
        self.drawing_area.change_color(col)

    def random_word(self):
        workbook = load_workbook('./img/word.xlsx')
        sheet = workbook['Sheet1']
        ran_num = random.randint(1,3404)
        data = sheet.cell(row=ran_num,column=1).value
        self.answer_text_edit.setText(data)

    def register_question(self):
        ip = socket.gethostbyname(socket.gethostname())

        answer = self.answer_text_edit.toPlainText()
        if answer:
            query = "SELECT MAX(QIndex) FROM question"
            self.MariaDB.execute(query)
            result = self.MariaDB.fetchone()
            max_index = result[0] if result[0] else 0
            index = max_index + 1
            image_name = '%s.jpg'%index
            self.save_img(image_name)
            self.upload_ftp(image_name)
            query = "INSERT INTO question (QIndex, username, ans, length, img_path, ans_list) VALUES (%d, '%s', '%s', %d, '%s', '[]')"%(index,ip,answer,len(answer),image_name)
            self.MariaDB.execute(query)

            # 리스트 뷰 새로고침
            self.parent().load_data_to_list_view()

        self.close()

    def save_img(self,img_name):
        pixmap = QPixmap(self.drawing_area.size())
        self.drawing_area.render(pixmap)
        pixmap.save('./img/%s'%img_name)

    def upload_ftp(self,img_name):
        with ftplib.FTP() as ftp:
            ftp.connect(host='host', port=25904)
            ftp.encoding = 'utf-8'
            username = 'user'
            password = 'pwd'
            filename = img_name
            ftp.login(user=username, passwd=password)
            ftp.cwd('/Test/img')

            with open(file=r'./img/%s'%img_name, mode='rb') as wf:
                ftp.storbinary(f'STOR %s'%img_name,wf)
            ftp.quit()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    main_ui = MainUI()
    main_ui.show()
    sys.exit(app.exec_())
